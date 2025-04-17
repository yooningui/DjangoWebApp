import logging
import string
from time import sleep
import unicodedata
from django.test import TestCase
import requests
import random
import json
import copy
import os
import subprocess
import signal
import time
from datetime import datetime
import threading
import coverage
from home.apps import HomeConfig
# Make sure openpyxl is installed: pip install openpyxl
from openpyxl import Workbook, load_workbook

logging.basicConfig(
    filename='app.log',  # Logs will be written to app.log in the current directory
    level=logging.DEBUG,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

interesting_logger = logging.getLogger("interesting")
ih = logging.FileHandler("interesting.log")
ih.setFormatter(logging.Formatter("%(asctime)s %(message)s", datefmt="%Y-%m-%d %H:%M:%S"))
interesting_logger.addHandler(ih)
interesting_logger.setLevel(logging.INFO)

cov = coverage.Coverage()
cov.start()

def start_server(p):
    if p is not None:
        os.kill(p.pid, signal.CTRL_C_EVENT)  # FOR WINDOWS
    command = ["py", "manage.py", "runserver"]
    try:
        with open("server_output.txt", "a") as out_file, open("server_error.txt", "a") as err_file:
            p = subprocess.Popen(
                command,
                shell=True,  # windows
                creationflags=subprocess.CREATE_NEW_PROCESS_GROUP,  # windows
                stdout=out_file, 
                stderr=err_file
            )
        logger.info("Server started")
        sleep(1)
        return p
    except Exception as e:
        logger.error("Error starting server: %s", e)

class FuzzingTestCase(TestCase):
    def setUp(self):
        self.app_name = HomeConfig.name
        self.base_url = 'http://127.0.0.1:8000/datatb/product/'
        self.endpoint_url = 'add/'
        self.url = self.base_url + self.endpoint_url

        self.seedQ = []
        self.failureQ = []
        self.num_iterations = 20
        self.interesting = []
        self.testcases = 0
        self.coverage_before = None
        self.headers = {
            'Cookie': 'csrftoken=jr6DahhKuGKgXX6Dxb3F4iR9FgL; sessionid=bvlvh8bqcwhbzr2eqqk3b',
        }
        self.testCheckCov = []

        try:
            with open("seed.json", "r") as f:
                raw_seeds = json.load(f)
        except Exception as e:
            logger.error("Error loading seed.json: %s", e)
            raw_seeds = []
        self.seedQ = [{
            "seed": seed,
            "s": 0,        # Number of times chosen for fuzzing
            "f": 1,        # Fuzz count (start at 1 to avoid division by zero)
            "alpha": 20    # Base energy value; adjust as needed
        } for seed in raw_seeds]

    def test_fuzzing_request(self):
        signal.signal(signal.SIGINT, self.signal_handler)
        # Start server
        p = start_server(None)
        
        # Start thread to print interesting inputs vs. time
        stop_event = threading.Event()
        printer_thread = threading.Thread(target=self.print_interesting_vs_time, args=(self.interesting, stop_event))
        
        # Set fuzzing run timeout (e.g., 300 seconds)
        timeout_seconds = 300
        start_time = time.time()
        printer_thread.start()
        
        while time.time() - start_time < timeout_seconds:
            # Choose the next seed using our new strategy that prioritizes low (s, f) values
            seed_record = self.choose_next()
            # Compute the energy (number of mutations) using the adaptive power schedule
            energy = self.assign_energy(seed_record)
            logger.info("Processing seed: %s with energy: %s", seed_record["seed"], energy)
            
            for j in range(energy):
                mutated_seed = self.mutate_input(seed_record["seed"])
                # Increment the fuzz counter for the seed
                seed_record["f"] += 1
                
                # Build the request JSON with only required keys
                request_json = copy.deepcopy(mutated_seed)
                for key in list(request_json.keys()):
                    if key not in ("name", "info", "price"):
                        del request_json[key]
                        
                try:
                    logger.info("Requesting...")
                    self.testcases += 1
                    self.print_interesting_vs_testcases()
                    # Randomly choose a POST or GET request
                    if random.random() < 0.5:
                        output = requests.post(self.url, headers=self.headers, json=request_json, timeout=2)
                    else:
                        output = requests.get(self.url, params=request_json, timeout=2)
                    
                    # Check if the output reveals a bug
                    if self.reveals_bug(mutated_seed, output):
                        self.failureQ.append({
                            "seed": mutated_seed,
                            "request": "POST" if random.random() < 0.5 else "GET",
                            "response": output.text
                        })
                    elif self.is_interesting():
                        cov.start()
                        # If interesting, wrap it as a new seed record
                        new_seed_record = {
                            "seed": mutated_seed,
                            "s": 0,
                            "f": 1,
                            "alpha": 20
                        }
                        self.seedQ.append(new_seed_record)
                        self.interesting.append(mutated_seed)
                        interesting_logger.info(json.dumps(mutated_seed))
                except requests.exceptions.RequestException as e:
                    self.failureQ.append({
                        "seed": mutated_seed,
                        "request": "UNKNOWN",
                        "response": f"Request failed: {e}"
                    })
                    p = start_server(p)
            
            logger.info("Current seed queue: %s", [record["seed"] for record in self.seedQ])
            # self.write_interesting_to_excel(self.interesting, "interesting_inputs.xlsx")
        
        stop_event.set()
        printer_thread.join()
        logger.info("Test Coverage Checks: %s", self.testCheckCov)
        cov.stop()
        cov.save()
        cov.report()
        
        # Write the updated seed and failure queues to file
        with open("seed.json", "w") as f:
            # Write only the seed parts
            json.dump([record["seed"] for record in self.seedQ], f)
        with open("failure.json", "w") as f:
            json.dump(self.failureQ, f)

    def mutate_input(self, input_data):
        logger.debug("Input data: %s", input_data)
        mutations = (
            "bitflip", "byteflip", "arith inc/dec", "interesting values",
            "random bytes", "delete bytes", "insert bytes", "overwrite bytes", "cross over"
        )
        mutation_chose = random.choice(mutations)
        mutated_data = {}
        for key, value in input_data.items():
            if key in ("name", "info", "price"):
                mutated_data[key] = self.apply_mutation(value, mutation_chose, key)
        logger.debug("Mutated data: %s", mutated_data)
        return mutated_data

    def apply_mutation(self, data, mutation, key):
        if not isinstance(data, str):
            data = str(data)
        if data == "" or data is None:
            data = self.random_string(key)
        mutated_input = bytearray()
        mutated_input.extend(data.encode("ascii"))
        match mutation:
            case "bitflip":
                n = random.choice((1, 2, 4))
                for i in range(0, len(mutated_input) * 8):
                    byte_index = i // 8
                    bit_index = i % 8
                    for _ in range(n):
                        if i < len(mutated_input) * 8:
                            mutated_input[byte_index] ^= (1 << bit_index)
            case "byteflip":
                n = random.choice((1, 2, 4))
                for i in range(len(mutated_input)):
                    for j in range(n):
                        if (i + j) < len(mutated_input):
                            mutated_input[i + j] ^= 0xFF
            case "arith inc/dec":
                n = random.choice((1, 2, 4))
                operator = random.choice((1, -1))
                for i in range(len(mutated_input)):
                    for j in range(n):
                        if (i + j) < len(mutated_input):
                            mutated_input[i + j] = (mutated_input[i + j] + operator) % 256
            case "interesting values":
                interesting_values = (0x5c, 0x00, 0xFF, 0x7F, 0x80, 0x01, 0x7E, 0x7D, 0x7C,
                                      0x02, 0x03, 0x04, 0x05, 0x06, 0x07, 0x08, 0x09, 0x0A,
                                      0x0B, 0x0C, 0x0D, 0x0E, 0x0F)
                n = random.choice((1, 2, 4))
                for i in range(len(mutated_input)):
                    for j in range(n):
                        if (i + j) < len(mutated_input):
                            mutated_input[i + j] = random.choice(interesting_values)
            case "random bytes":
                byte_index = random.randint(0, len(mutated_input) - 1)
                mutated_input[byte_index] = random.randint(0, 255)
            case "delete bytes":
                size = random.randint(1, 4)
                start = random.randint(0, len(mutated_input))
                del mutated_input[start:start + size]
            case "insert bytes":
                size = random.randint(1, 4)
                start = random.randint(0, len(mutated_input))
                mutated_input[start:start] = bytearray(random.choices(range(256), k=size))
            case "overwrite bytes":
                size = random.randint(1, 4)
                start = random.randint(0, len(mutated_input))
                mutated_input[start:start + size] = bytearray(random.choices(range(256), k=size))
            case "cross over":
                if self.seedQ:
                    data2 = random.choice(self.seedQ)["seed"]
                    other_data = bytearray()
                    other_data.extend(str(data2.get(key, "")).encode("ascii"))
                    if len(other_data) < len(mutated_input):
                        splice_loc = random.randint(0, len(other_data))
                    else:
                        splice_loc = random.randint(0, len(mutated_input))
                    mutated_input[splice_loc:] = other_data[splice_loc:]
        try:
            mutated_str = bytes(mutated_input).decode("ascii", errors="ignore")
        except Exception as ex:
            logger.error("Error decoding mutated input: %s", ex)
            mutated_str = self.random_string(key)
        mutated_str = unicodedata.normalize("NFKD", mutated_str)
        if mutated_str == "" or mutated_str is None:
            mutated_str = self.random_string(key)
        return mutated_str

    def random_string(self, key):
        if key == "name":
            data = ''.join(random.choice(string.printable) for i in range(128))
        elif key == "info":
            data = ''.join(random.choice(string.printable) for i in range(1500))
        elif key == "price":
            data = str(random.randint(0, 1000))
        return data

    def choose_next(self):
        if not self.seedQ:
            raise Exception("Seed queue is empty!")
        # Select the seed record with the smallest (s, f) values,
        # prioritizing those that have been fuzzed less often.
        chosen_record = sorted(self.seedQ, key=lambda rec: (rec["s"], rec["f"]))[0]
        # Increment the counter for the number of times chosen (s)
        chosen_record["s"] += 1
        return chosen_record
    
    def assign_energy(self, seed):
        MAX_ENERGY = 64  # Maximum energy per fuzzing iteration
        # Compute energy using an exponential schedule:
        # energy = (alpha * 2^(s)) / f, capped at MAX_ENERGY.
        energy = (seed["alpha"] * (2 ** seed["s"])) / seed["f"]
        energy = min(energy, MAX_ENERGY)
        # Ensure at least one mutation is attempted.
        return int(max(1, energy))

    
    def reveals_bug(self, seed, output):
        if output.status_code == 200:
            logger.info("Request successful! Request: %s", seed)
            logger.debug("Response: %s", output.text)
            logger.debug("-" * 60)
            return False
        else:
            logger.error("Request failed: Status code %s", output.status_code)
            logger.error("Seed: %s", seed)
            logger.error("Response: %s", output.text)
            logger.error("-" * 60)
            return True

    def is_interesting(self):
        cov.stop()
        cov.save()
        if self.coverage_before is None:
            self.coverage_before = 0
            return True

        # Note: Ensure that you extract a numeric value for comparison.
        # Here, we assume cov.json_report returns a numeric metric (this may need adjustment).
        coverage_after = cov.json_report(pretty_print=True)
        if coverage_after > self.coverage_before:
            self.testCheckCov.append(f"Coverage increased from {self.coverage_before} to {coverage_after}")
            self.coverage_before = coverage_after
            return True
        return False
    
    def signal_handler(self, sig, frame):
        logger.info("Exiting...")
        exit(0)

    def print_interesting_vs_time(self, list_data, stop_event):
        # Log header once
        header = (
            "Timestamp".ljust(25)
            + "No. of interesting".ljust(20)
            + "Recent Interesting Inputs"
        )
        logger.info(header)
        while not stop_event.is_set():
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            count = len(list_data)
            # Show only the last three interesting inputs for brevity
            recent_interesting = list_data[-3:]
            recent_interesting_str = ", ".join(json.dumps(inp) for inp in recent_interesting)
            line = timestamp.ljust(25) + str(count).ljust(20) + recent_interesting_str
            logger.info(line)
            time.sleep(1)

    def print_interesting_vs_testcases(self):
        # Log current testcase count and number of interesting cases
        logger.info("Testcases: %s, Interesting: %s", self.testcases, len(self.interesting))

    # def write_interesting_to_excel(self, interesting_list, filename):
    #     """Append the interesting inputs to an Excel file if it exists, or create a new one."""
        
    #     if os.path.exists(filename):
    #         wb = load_workbook(filename)
    #         if "Interesting Inputs" in wb.sheetnames:
    #             ws = wb["Interesting Inputs"]
    #         else:
    #             ws = wb.active
    #             ws.title = "Interesting Inputs"
    #     else:
    #         wb = Workbook()
    #         ws = wb.active
    #         ws.title = "Interesting Inputs"
    #         ws.append(["Name", "Info", "Price"])
        
    #     for inp in interesting_list:
    #         name = inp.get("name", "")
    #         info = inp.get("info", "")
    #         price = inp.get("price", "")
    #         ws.append([name, info, price])
        
    #     wb.save(filename)
    #     logger.info("Written %s interesting inputs to %s", len(interesting_list), filename)

    def test_regexbomb(self):
        """
        Test using a regex bomb input.
        This input is crafted to trigger catastrophic backtracking if processed by a vulnerable regex.
        """
        regex_bomb = "((a+)+)+"
        request_json = {"name": regex_bomb, "info": regex_bomb, "price": "0"}
        try:
            logger.info("Starting regexbomb test...")
            output = requests.post(self.url, headers=self.headers, json=request_json, timeout=10)
            logger.info("Regexbomb test response: %s", output.text)
        except Exception as e:
            logger.error("Regexbomb test encountered an error: %s", e)

    def test_out_of_memory(self):
        """
        Test sending a huge payload to simulate out-of-memory conditions.
        Adjust the payload size as needed for your environment.
        """
        huge_info = "A" * (10 ** 7)  # 10 million characters; be cautious with size
        request_json = {"name": "OOM Test", "info": huge_info, "price": "100"}
        try:
            logger.info("Starting out-of-memory test...")
            output = requests.post(self.url, headers=self.headers, json=request_json, timeout=10)
            logger.info("Out-of-memory test response: %s", output.text)
        except Exception as e:
            logger.error("Out-of-memory test encountered an error: %s", e)

    def test_race_condition(self):
        """
        Test to simulate a race condition by firing multiple concurrent requests.
        """
        def send_request(i):
            req_json = {"name": f"RaceTest {i}", "info": "Concurrent access", "price": str(123)}
            try:
                response = requests.post(self.url, headers=self.headers, json=req_json, timeout=5)
                logger.info("Race condition request %s response: %s", i, response.text)
            except Exception as e:
                logger.error("Race condition request %s failed: %s", i, e)

        threads = []
        for i in range(20):  # Launch 20 concurrent requests
            t = threading.Thread(target=send_request, args=(i,))
            threads.append(t)
            t.start()
        for t in threads:
            t.join()
        logger.info("Race condition test completed.")
